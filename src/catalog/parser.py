"""Parse Incidentes.xlsx into a list of IncidentTemplate objects."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from src.models import Category, IncidentTemplate, Severity, TicketType

_SEVERITY_MAP: dict[str, Severity] = {
    "baja": Severity.LOW,
    "media": Severity.MEDIUM,
    "alta": Severity.HIGH,
    "crítica": Severity.CRITICAL,
}

_CATEGORY_MAP: dict[str, Category] = {
    "terminales / pos": Category.POS,
    "terminales/pos": Category.POS,
    "impresoras / tickets": Category.IMP,
    "impresoras/tickets": Category.IMP,
    "internet / conectividad": Category.NET,
    "internet/conectividad": Category.NET,
    "electricidad / energía": Category.ELE,
    "electricidad / energia": Category.ELE,
    "electricidad/energía": Category.ELE,
    "equipos de cómputo": Category.EQU,
    "equipos de computo": Category.EQU,
    "local / infraestructura": Category.INF,
    "local/infraestructura": Category.INF,
    "materiales / suministros": Category.MAT,
    "materiales/suministros": Category.MAT,
    "operación de ventas": Category.VEN,
    "operacion de ventas": Category.VEN,
    "pagos y premios": Category.PAG,
    "contabilidad / cuadres": Category.CON,
    "contabilidad/cuadres": Category.CON,
    "seguridad / fraude": Category.FRA,
    "seguridad/fraude": Category.FRA,
    "reclamos de clientes": Category.REC,
}

_TICKET_TYPE_MAP: dict[str, TicketType] = {
    "incidente": TicketType.INCIDENTE,
    "alerta": TicketType.ALERTA,
    "reclamo": TicketType.RECLAMO,
}


def _parse_severity(raw: str) -> Severity:
    normalized = raw.strip().lower().split("(")[0].strip()
    for key, sev in _SEVERITY_MAP.items():
        if key in normalized:
            return sev
    return Severity.MEDIUM


def _resolve_category(raw: str) -> Category:
    normalized = raw.strip().lower()
    for key, cat in _CATEGORY_MAP.items():
        if key in normalized:
            return cat
    raise ValueError(f"Unknown category: {raw!r}")


def parse_catalog(catalog_path: Path | str) -> list[IncidentTemplate]:
    """Read the Excel catalog and return structured templates."""
    wb = load_workbook(str(catalog_path), read_only=True, data_only=True)
    ws = wb.active

    templates: list[IncidentTemplate] = []
    counters: dict[str, int] = defaultdict(int)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 7:
            continue
        cat_raw, sub_cat, ticket_type_raw, description, severity_raw, sla, requires_image = row[:7]

        if not cat_raw or not description:
            continue

        category = _resolve_category(str(cat_raw))
        counters[category.value] += 1
        code = f"{category.value}-{counters[category.value]:03d}"

        ticket_type = _TICKET_TYPE_MAP.get(
            str(ticket_type_raw or "").strip().lower(),
            TicketType.INCIDENTE,
        )

        templates.append(IncidentTemplate(
            code=code,
            category=category,
            sub_category=str(sub_cat or "").strip(),
            name=str(description).strip(),
            description=str(description).strip(),
            severity=_parse_severity(str(severity_raw or "Media")),
            ticket_type=ticket_type,
            sla=str(sla or "").strip(),
            requires_image=bool(requires_image and str(requires_image).strip().lower() in ("sí", "si", "yes", "true", "1")),
        ))

    wb.close()
    return templates


def load_catalog_text(catalog_path: Path | str) -> str:
    """Return a readable text version of the Excel catalog for LLM prompts."""
    templates = parse_catalog(catalog_path)
    lines: list[str] = ["# Catálogo de Incidentes — Agencias de Lotería\n"]

    current_cat = None
    for t in templates:
        if t.category != current_cat:
            current_cat = t.category
            cat_names = {
                Category.POS: "Terminales / POS",
                Category.IMP: "Impresoras / Tickets",
                Category.NET: "Internet / Conectividad",
                Category.ELE: "Electricidad / Energía",
                Category.EQU: "Equipos de Cómputo",
                Category.INF: "Local / Infraestructura",
                Category.MAT: "Materiales / Suministros",
                Category.VEN: "Operación de Ventas",
                Category.PAG: "Pagos y Premios",
                Category.CON: "Contabilidad / Cuadres",
                Category.FRA: "Seguridad / Fraude",
                Category.REC: "Reclamos de Clientes",
            }
            lines.append(f"\n## {cat_names.get(current_cat, current_cat.value)}\n")

        lines.append(f"### {t.code} – {t.name}")
        lines.append(f"- **Subcategoría:** {t.sub_category}")
        lines.append(f"- **Tipo de ticket:** {t.ticket_type.value}")
        lines.append(f"- **Severidad:** {t.severity.value}")
        lines.append(f"- **SLA:** {t.sla}")
        lines.append("")

    return "\n".join(lines)
